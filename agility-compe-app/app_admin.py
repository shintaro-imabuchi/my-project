import io
import unicodedata

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import streamlit as st

from supabase_client import get_supabase
from utils.settings import (
    get_registration_open,
    set_registration_open,
    get_event_fees,
    set_event_fees,
    get_login_message,
    set_login_message,
    get_home_info_message,
    set_home_info_message,
    get_guideline_url,
    set_guideline_url,
    get_notice_url,
    set_notice_url,
    get_top_image_url,
    set_top_image_url,
)


def calc_fee(events: list[str], event_fees: dict[str, int] | None = None) -> int:
    """参加種目リストから参加料金の合計を計算する。"""
    if event_fees is None:
        event_fees = get_event_fees()
    return sum(event_fees.get(e, 0) for e in events)


def check_admin_password() -> bool:
    """管理者パスワードを確認する。

    認証済みの場合はTrueを返す。未認証の場合はログインフォームを表示してFalseを返す。
    """
    if st.session_state.get("admin_authenticated"):
        return True

    msg = get_login_message()
    if msg:
        st.markdown(f"### **{msg}**")

    st.subheader("管理者ログイン")
    password = st.text_input("パスワード", type="password", key="admin_password_input")

    if st.button("ログイン", type="primary", use_container_width=True):
        if password == st.secrets["admin"]["password"]:
            st.session_state["admin_authenticated"] = True
            st.rerun()
        else:
            st.error("パスワードが間違っています。")

    return False


def fetch_participants() -> list[dict] | None:
    """参加者と犬情報の一覧をSupabaseから取得する。"""
    try:
        response = get_supabase().rpc("get_participants_with_dogs").execute()
        return response.data
    except Exception as e:
        st.error(f"データの取得に失敗しました: {e}")
        return None


def _str_width(s: str) -> int:
    """文字列の表示幅を返す（全角=2、半角=1）。"""
    return sum(2 if unicodedata.east_asian_width(c) in ("W", "F") else 1 for c in str(s))


def generate_excel(data: list[dict]) -> bytes:
    """参加者一覧をExcel(.xlsx)形式で生成してバイト列で返す。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "参加者一覧"

    headers = list(data[0].keys())
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    total_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")

    # ヘッダー行
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    fee_col = headers.index("参加料金") + 1

    # データ行
    for row_idx, row in enumerate(data, start=2):
        is_total = row.get("参加者名") == "合計"
        for col_idx, key in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
            if col_idx == fee_col:
                cell.alignment = Alignment(horizontal="right")
            if is_total:
                cell.font = Font(bold=True)
                cell.fill = total_fill

    # 列幅を最大文字列幅に合わせる
    for col_idx, key in enumerate(headers, start=1):
        max_width = _str_width(key)
        for row in data:
            max_width = max(max_width, _str_width(row[key]))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_width + 2

    # A4縦・1ページ幅に収める印刷設定
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def fetch_user_emails() -> list[dict] | None:
    """ユーザーの氏名とメールアドレスの一覧をSupabaseから取得する。"""
    try:
        response = get_supabase().rpc("get_user_emails").execute()
        return response.data
    except Exception as e:
        st.error(f"メールアドレスの取得に失敗しました: {e}")
        return None


def show_bcc_list() -> None:
    """BCC用メールアドレス一覧を表示する。"""
    st.markdown("#### BCC用メールアドレス一覧")
    with st.spinner("読み込み中..."):
        users = fetch_user_emails()
    if users is None:
        return
    if not users:
        st.info("登録済みユーザーがいません。")
        return

    fmt = st.radio(
        "書式",
        options=["引用符付き氏名 + メールアドレス", "メールアドレスのみ"],
        horizontal=True,
        key="bcc_format",
    )
    if fmt == "引用符付き氏名 + メールアドレス":
        bcc = ",\n".join(
            f'"{u["user_name"]}" <{u["email"]}>' for u in users if u.get("email")
        )
    else:
        bcc = ",\n".join(u["email"] for u in users if u.get("email"))

    st.text_area("BCCにコピー＆ペーストしてください", value=bcc, height=150)


def show_event_settings() -> None:
    """競技種目と参加料金の設定UIを表示する。

    行の追加・編集・削除ができる。保存するとSupabaseに反映される。
    並び順を変えたい場合は削除してから再追加する。
    """
    st.markdown("#### 競技種目設定")
    st.caption("種目名・料金を編集し「保存」してください。並び順を変えたい場合は削除して再追加してください。")

    if "event_list_edit" not in st.session_state:
        event_fees = get_event_fees()
        st.session_state["event_list_edit"] = [
            {"name": n, "fee": f} for n, f in event_fees.items()
        ]

    events: list[dict] = st.session_state["event_list_edit"]
    n = len(events)

    def _read_inputs() -> None:
        """ウィジェットの現在値を events リストに反映する。"""
        for j in range(n):
            events[j]["name"] = str(st.session_state.get(f"ev_name_{j}", events[j]["name"]))
            events[j]["fee"] = int(st.session_state.get(f"ev_fee_{j}", events[j]["fee"]) or 0)

    def _clear_inputs() -> None:
        """ウィジェットの session_state キーを削除し、次回レンダリングで value= から再初期化させる。"""
        for j in range(n + 1):
            st.session_state.pop(f"ev_name_{j}", None)
            st.session_state.pop(f"ev_fee_{j}", None)

    # ヘッダー行
    h = st.columns([3, 2, 1])
    h[0].markdown("**種目名**")
    h[1].markdown("**参加料金（円）**")

    for i, event in enumerate(events):
        c = st.columns([3, 2, 1])
        c[0].text_input(
            "種目名", value=event["name"], key=f"ev_name_{i}", label_visibility="collapsed"
        )
        c[1].number_input(
            "料金", value=event["fee"], min_value=0, step=100, format="%d",
            key=f"ev_fee_{i}", label_visibility="collapsed",
        )
        if c[2].button("削除", key=f"ev_del_{i}", use_container_width=True):
            _read_inputs()
            events.pop(i)
            _clear_inputs()
            st.rerun()

    col_add, col_reset, col_save = st.columns(3)

    if col_add.button("＋ 種目を追加", use_container_width=True):
        _read_inputs()
        events.append({"name": "", "fee": 0})
        _clear_inputs()
        st.rerun()

    if col_reset.button("リセット（保存済みに戻す）", use_container_width=True, key="reset_event_fees"):
        st.session_state.pop("event_list_edit", None)
        _clear_inputs()
        st.rerun()

    if col_save.button("保存", type="primary", use_container_width=True, key="save_event_fees"):
        _read_inputs()
        new_fees: dict[str, int] = {
            ev["name"].strip(): ev["fee"]
            for ev in events
            if ev["name"].strip()
        }
        if not new_fees:
            st.error("種目を1つ以上入力してください。")
            return
        set_event_fees(new_fees)
        st.session_state["event_list_edit"] = [{"name": n, "fee": f} for n, f in new_fees.items()]
        _clear_inputs()
        st.success(f"種目リストを保存しました（{len(new_fees)}種目）。")
        st.rerun()


def show_login_message_settings() -> None:
    """ログイン画面開催名称設定UIを表示する。"""
    st.markdown("#### ログイン画面開催名称設定")
    st.caption("ログイン画面の先頭に表示する開催名称を設定してください。空白にすると非表示になります。")
    current = get_login_message()
    new_msg = st.text_input(
        "開催名称（全角20文字程度）",
        value=current or "",
        max_chars=60,
        key="login_message_input",
    )
    if st.button("保存", type="primary", key="save_login_message"):
        set_login_message(new_msg)
        st.success("開催名称を保存しました。")
        st.rerun()


def show_link_url_settings() -> None:
    """リンクボタンURL設定UIを表示する。"""
    st.markdown("#### リンクURL設定")
    st.caption("空白にするとボタンが非表示になります。")

    guideline_url = st.text_input(
        "練習会要綱を見る（トップ画面）",
        value=get_guideline_url() or "",
        key="guideline_url_input",
    )
    notice_url = st.text_input(
        "練習会実施のご案内を見る（ホーム画面）",
        value=get_notice_url() or "",
        key="notice_url_input",
    )
    if st.button("保存", type="primary", key="save_link_urls"):
        set_guideline_url(guideline_url)
        set_notice_url(notice_url)
        st.success("URLを保存しました。")
        st.rerun()


def show_top_image_settings() -> None:
    """トップ画面表示画像設定UIを表示する。"""
    st.markdown("#### トップ画面画像設定")
    st.caption("トップ画面最下部に表示する画像を設定します。Supabase Storageの「images」バケット（パブリック）が必要です。")

    current_url = get_top_image_url()
    if current_url:
        st.image(current_url, caption="現在の画像")
        if st.button("画像を削除", key="delete_top_image"):
            try:
                get_supabase().storage.from_("images").remove(["top_image"])
            except Exception:
                pass
            set_top_image_url(None)
            st.success("画像を削除しました。")
            st.rerun()
    else:
        st.info("画像は設定されていません。")

    uploaded_file = st.file_uploader(
        "画像ファイルを選択",
        type=["jpg", "jpeg", "png", "gif", "webp"],
        key="top_image_upload",
    )
    if uploaded_file is not None:
        if st.button("アップロード", type="primary", key="upload_top_image"):
            try:
                file_bytes = uploaded_file.read()
                mime_type = uploaded_file.type
                bucket = get_supabase().storage.from_("images")
                bucket.upload(
                    path="top_image",
                    file=file_bytes,
                    file_options={"content-type": mime_type, "upsert": "true"},
                )
                url = bucket.get_public_url("top_image")
                set_top_image_url(url)
                st.success("画像をアップロードしました。")
                st.rerun()
            except Exception as e:
                st.error(f"アップロードに失敗しました: {e}")


def show_home_info_message_settings() -> None:
    """ログイン後ホーム画面のお知らせ設定UIを表示する。"""
    st.markdown("#### ホーム画面お知らせ設定")
    st.caption("ログイン後のホーム画面に表示するお知らせを設定してください。空白にすると非表示になります。")
    current = get_home_info_message()
    new_msg = st.text_area(
        "お知らせ本文",
        value=current or "",
        height=120,
        key="home_info_message_input",
    )
    if st.button("保存", type="primary", key="save_home_info_message"):
        set_home_info_message(new_msg)
        st.success("お知らせを保存しました。")
        st.rerun()


def show_admin_home() -> None:
    """管理者ホーム画面を表示する。"""
    st.subheader("管理者メニュー")

    show_login_message_settings()

    st.divider()

    show_home_info_message_settings()

    st.divider()

    show_link_url_settings()

    st.divider()

    show_top_image_settings()

    st.divider()

    show_event_settings()

    st.divider()

    st.markdown("#### 新規登録受付設定")
    registration_open = get_registration_open()
    new_value = st.toggle("新規登録を受け付ける", value=registration_open)

    if new_value != registration_open:
        set_registration_open(new_value)
        status = "受付中" if new_value else "締め切り"
        st.success(f"新規登録を「{status}」に変更しました。")
        st.rerun()

    st.divider()

    event_fees = get_event_fees()

    with st.spinner("読み込み中..."):
        participants = fetch_participants()

    if participants is not None:
        st.markdown("#### 参加者・犬情報一覧")

        # 画面表示・Excelダウンロード共通のデータを構築
        display_data: list[dict] = []
        for i, row in enumerate(participants, start=1):
            events = row.get("events") or []
            display_data.append({
                "No.": i,
                "参加者名": row["user_name"],
                "犬名": row["dog_name"],
                "犬種": row["breed"],
                "クラス": row["dog_class"],
                "参加種目": "、".join(events),
                "参加料金": f"{calc_fee(events, event_fees):,}円",
            })
        total_fee = sum(event_fees.get(e, 0) for row in participants for e in (row.get("events") or []))
        display_data.append({
            "No.": "", "参加者名": "合計", "犬名": "", "犬種": "", "クラス": "", "参加種目": "",
            "参加料金": f"{total_fee:,}円",
        })

        col1, col2 = st.columns(2)
        if col1.button("画面表示", use_container_width=True):
            st.session_state["show_participants"] = True
        col2.download_button(
            label="Excelダウンロード",
            data=generate_excel(display_data),
            file_name="参加者一覧.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        if st.session_state.get("show_participants"):
            if not participants:
                st.info("登録データがありません。")
            else:
                st.dataframe(
                    display_data,
                    hide_index=True,
                    use_container_width=True,
                    column_config={"参加料金": st.column_config.TextColumn(width="small")},
                )


    st.divider()
    show_bcc_list()

    st.divider()
    if st.button("ログアウト", use_container_width=True):
        del st.session_state["admin_authenticated"]
        st.rerun()


def main() -> None:
    """管理者アプリのメインエントリーポイント。"""
    st.set_page_config(
        page_title="管理者 | アジリティー練習会",
        layout="wide",
    )

    if not check_admin_password():
        return

    show_admin_home()


if __name__ == "__main__":
    main()
